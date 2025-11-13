from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.user import User
from pydantic import BaseModel, EmailStr
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime
import secrets
import hashlib

router = APIRouter()

def generate_salt() -> str:
    """Generate a random salt"""
    return secrets.token_hex(16)

def hash_password_with_salt(password: str, salt: str) -> str:
    """Hash password with salt using SHA256"""
    return hashlib.sha256((password + salt).encode()).hexdigest()

# -----------------------------
# Pydantic Schemas
# -----------------------------
class UserUpdate(BaseModel):
    name: Optional[str] = None
    department: Optional[str] = None
    designation: Optional[str] = None
    hod: Optional[str] = None
    supervisor: Optional[str] = None
    email: Optional[EmailStr] = None
    password: Optional[str] = None

@router.get("/")
def get_all_users(db: Session = Depends(get_db)):
    """Get all users"""
    users = db.query(User).order_by(User.id).all()
    
    return [
        {
            "id": user.id,
            "emp_id": user.emp_id,
            "name": user.name,
            "department": user.department,
            "designation": user.designation,
            "hod": user.hod,
            "supervisor": user.supervisor,
            "email": user.email,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "updated_at": user.updated_at.isoformat() if user.updated_at else None
        }
        for user in users
    ]

@router.get("/export-to-excel")
def export_to_excel(db: Session = Depends(get_db)):
    """
    Export all users to Excel file with all columns from users table.
    """
    # Query all users
    users = db.query(User).order_by(User.id).all()
    
    if not users:
        raise HTTPException(status_code=404, detail="No users found to export")
    
    # Create a new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    # Define column headers (all columns from users table)
    headers = [
        "ID",
        "Employee ID",
        "Name",
        "Department",
        "Designation",
        "HOD",
        "Supervisor",
        "Email",
        "Created At",
        "Updated At"
    ]
    
    # Style the header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    for row_num, user in enumerate(users, 2):
        ws.cell(row=row_num, column=1, value=user.id)
        ws.cell(row=row_num, column=2, value=user.emp_id)
        ws.cell(row=row_num, column=3, value=user.name)
        ws.cell(row=row_num, column=4, value=user.department)
        ws.cell(row=row_num, column=5, value=user.designation)
        ws.cell(row=row_num, column=6, value=user.hod)
        ws.cell(row=row_num, column=7, value=user.supervisor)
        ws.cell(row=row_num, column=8, value=user.email)
        ws.cell(row=row_num, column=9, value=user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else None)
        ws.cell(row=row_num, column=10, value=user.updated_at.strftime("%Y-%m-%d %H:%M:%S") if user.updated_at else None)
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = ws.cell(row=1, column=col_num).column_letter
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            if row[0].value:
                max_length = max(max_length, len(str(row[0].value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Set row height for header
    ws.row_dimensions[1].height = 25
    
    # Save workbook to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"users_export_{timestamp}.xlsx"
    
    # Return as downloadable file
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

@router.get("/{emp_id}")
def get_user_by_emp_id(emp_id: int, db: Session = Depends(get_db)):
    """Get user by Employee ID"""
    user = db.query(User).filter(User.emp_id == emp_id).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {
        "id": user.id,
        "emp_id": user.emp_id,
        "name": user.name,
        "department": user.department,
        "designation": user.designation,
        "hod": user.hod,
        "supervisor": user.supervisor,
        "email": user.email,
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "updated_at": user.updated_at.isoformat() if user.updated_at else None
    }

@router.put("/{emp_id}")
def update_user(emp_id: int, data: UserUpdate, db: Session = Depends(get_db)):
    """Update user information"""
    user = db.query(User).filter(User.emp_id == emp_id).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update fields if provided
    if data.name is not None:
        user.name = data.name
    
    if data.department is not None:
        user.department = data.department
    
    if data.designation is not None:
        user.designation = data.designation
    
    if data.hod is not None:
        user.hod = data.hod
    
    if data.supervisor is not None:
        user.supervisor = data.supervisor
    
    if data.email is not None:
        # Check if new email already exists
        existing = db.query(User).filter(
            User.email == data.email,
            User.emp_id != emp_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Email already exists")
        user.email = data.email
    
    if data.password is not None:
        # Generate new salt and hash password
        password_salt = generate_salt()
        password_hash = hash_password_with_salt(data.password, password_salt)
        user.password_hash = password_hash
        user.password_salt = password_salt
    
    db.commit()
    db.refresh(user)
    
    return {
        "message": "User updated successfully",
        "id": user.id,
        "emp_id": user.emp_id,
        "email": user.email
    }

@router.delete("/{emp_id}")
def delete_user(emp_id: int, db: Session = Depends(get_db)):
    """Delete user"""
    user = db.query(User).filter(User.emp_id == emp_id).first()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    db.delete(user)
    db.commit()
    
    return {
        "message": "User deleted successfully"
    }

