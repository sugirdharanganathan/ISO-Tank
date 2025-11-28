from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.tank_header import Tank
from app.models.tank_details import TankDetails
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime

router = APIRouter()

# --- Helper function to convert "" to None for numeric types ---
def to_float_or_none(value: any) -> float | None:
    if value == "" or value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail=f"Invalid numeric value: {value}")

@router.post("/")
def create_tank(data: dict, db: Session = Depends(get_db)):
    required_fields = [
        "tank_number", "mfgr", "pv_code", "un_iso_code",
        "capacity_l", "mawp", "design_temperature", "tare_weight_kg",
        "mgw_kg", "mpl_kg", "size", "pump_type",
        "vesmat", "gross_kg", "net_kg", "color_body_frame"
    ]

    for field in required_fields:
        if field not in data or data[field] == "":
            raise HTTPException(status_code=400, detail=f"Required field '{field}' is missing or empty")

    existing_tank = db.query(Tank).filter(Tank.tank_number == data["tank_number"]).first()
    if existing_tank:
        raise HTTPException(status_code=400, detail=f"Tank number '{data['tank_number']}' already exists")

    tank = Tank(
        tank_number=data["tank_number"],
        created_by=data.get("created_by")
    )
    db.add(tank)
    db.commit()
    db.refresh(tank)

    status = data.get("status", "active")
    if status not in ["active", "inactive"]:
        raise HTTPException(status_code=400, detail="status must be 'active' or 'inactive'")
    
    tank.status = status
    db.commit()
    db.refresh(tank)
    
    try:
        tank_detail = TankDetails(
            tank_id=tank.id,
            tank_number=data["tank_number"],
            status=status,
            mfgr=data["mfgr"],
            date_mfg=data.get("date_mfg") or None,
            pv_code=data["pv_code"],
            un_iso_code=data["un_iso_code"],
            capacity_l=to_float_or_none(data["capacity_l"]),
            mawp=to_float_or_none(data["mawp"]),
            design_temperature=data["design_temperature"],
            tare_weight_kg=to_float_or_none(data["tare_weight_kg"]),
            mgw_kg=to_float_or_none(data["mgw_kg"]),
            mpl_kg=to_float_or_none(data["mpl_kg"]),
            size=data["size"],
            pump_type=data["pump_type"],
            vesmat=data["vesmat"],
            gross_kg=to_float_or_none(data["gross_kg"]),
            net_kg=to_float_or_none(data["net_kg"]),
            color_body_frame=data["color_body_frame"],
            # --- NEW FIELDS FOR CREATION ---
            working_pressure=to_float_or_none(data.get("working_pressure")),
            cabinet_type=data.get("cabinet_type") or None,
            frame_type=data.get("frame_type") or None,
            # -------------------------------
            remark=data.get("remark") or None,
            lease=bool(data.get("lease", 0)),
            created_by=data.get("created_by") or None
        )

        db.add(tank_detail)
        db.commit()
        db.refresh(tank_detail)

    except Exception as e:
        db.rollback() 
        db.delete(tank) 
        db.commit()
        raise HTTPException(status_code=500, detail=f"Failed to create tank details: {str(e)}")

    return {
        "message": "Tank created successfully",
        "tank_id": tank.id,
        "data": {
            "tank_header": tank,
            "tank_details": tank_detail
        }
    }


@router.get("/")
def get_all_tanks(db: Session = Depends(get_db)):
    results = db.query(Tank, TankDetails).join(TankDetails, Tank.id == TankDetails.tank_id).all()
    
    return [
        {
            "id": r[0].id,
            "tank_number": r[0].tank_number,
            "status": r[1].status,
            "mfgr": r[1].mfgr,
            "date_mfg": r[1].date_mfg,
            "pv_code": r[1].pv_code,
            "un_iso_code": r[1].un_iso_code,
            "capacity_l": r[1].capacity_l,
            "mawp": r[1].mawp,
            "design_temperature": r[1].design_temperature,
            "tare_weight_kg": r[1].tare_weight_kg,
            "mgw_kg": r[1].mgw_kg,
            "mpl_kg": r[1].mpl_kg,
            "size": r[1].size,
            "pump_type": r[1].pump_type,
            "vesmat": r[1].vesmat,
            "gross_kg": r[1].gross_kg,
            "net_kg": r[1].net_kg,
            "color_body_frame": r[1].color_body_frame,
            # --- NEW FIELDS FOR GET ---
            "working_pressure": r[1].working_pressure,
            "cabinet_type": r[1].cabinet_type,
            "frame_type": r[1].frame_type,
            # --------------------------
            "remark": r[1].remark,                  
            "lease": int(r[1].lease),             
            "created_by": r[0].created_by
        }
        for r in results
    ]

@router.get("/export-to-excel")
def export_to_excel(db: Session = Depends(get_db)):
    # ... (This function is unchanged) ...
    results = db.query(Tank, TankDetails).join(TankDetails, Tank.id == TankDetails.tank_id).all()
    
    if not results:
        raise HTTPException(status_code=404, detail="No tank details found to export")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tank Details"
    
    headers = [
        "ID", "Tank ID", "Tank Number", "Status", "Manufacturer (MFGR)",
        "Date of Manufacture", "PV Code", "UN ISO Code", "Capacity (L)", "MAWP",
        "Design Temperature", "Tare Weight (kg)", "MGW (kg)", "MPL (kg)", "Size",
        "Pump Type", "VESMAT", "Gross (kg)", "Net (kg)", "Color Body Frame",
        "Remark", "Lease", "Created By", "Updated By"
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_num, (tank, tank_detail) in enumerate(results, 2):
        ws.cell(row=row_num, column=1, value=tank_detail.id)
        ws.cell(row=row_num, column=2, value=tank_detail.tank_id)
        ws.cell(row=row_num, column=3, value=tank_detail.tank_number or tank.tank_number)
        ws.cell(row=row_num, column=4, value=tank_detail.status)
        ws.cell(row=row_num, column=5, value=tank_detail.mfgr)
        ws.cell(row=row_num, column=6, value=tank_detail.date_mfg.strftime("%Y-%m-%d") if tank_detail.date_mfg else None)
        ws.cell(row=row_num, column=7, value=tank_detail.pv_code)
        ws.cell(row=row_num, column=8, value=tank_detail.un_iso_code)
        ws.cell(row=row_num, column=9, value=tank_detail.capacity_l)
        ws.cell(row=row_num, column=10, value=tank_detail.mawp)
        ws.cell(row=row_num, column=11, value=tank_detail.design_temperature)
        ws.cell(row=row_num, column=12, value=tank_detail.tare_weight_kg)
        ws.cell(row=row_num, column=13, value=tank_detail.mgw_kg)
        ws.cell(row=row_num, column=14, value=tank_detail.mpl_kg)
        ws.cell(row=row_num, column=15, value=tank_detail.size)
        ws.cell(row=row_num, column=16, value=tank_detail.pump_type)
        ws.cell(row=row_num, column=17, value=tank_detail.vesmat)
        ws.cell(row=row_num, column=18, value=tank_detail.gross_kg)
        ws.cell(row=row_num, column=19, value=tank_detail.net_kg)
        ws.cell(row=row_num, column=20, value=tank_detail.color_body_frame)
        ws.cell(row=row_num, column=21, value=tank_detail.remark)
        ws.cell(row=row_num, column=22, value="Yes" if tank_detail.lease else "No")
        ws.cell(row=row_num, column=23, value=tank_detail.created_by)
        ws.cell(row=row_num, column=24, value=tank_detail.updated_by)
    
    for col_num, header in enumerate(headers, 1):
        column_letter = ws.cell(row=1, column=col_num).column_letter
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            if row[0].value:
                max_length = max(max_length, len(str(row[0].value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    ws.row_dimensions[1].height = 25
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tank_details_export_{timestamp}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.put("/{tank_id}")
def update_tank(tank_id: int, data: dict, db: Session = Depends(get_db)):
    tank = db.query(Tank).filter(Tank.id == tank_id).first()
    tank_detail = db.query(TankDetails).filter(TankDetails.tank_id == tank_id).first()

    if not tank or not tank_detail:
        raise HTTPException(status_code=404, detail="Tank not found")

    # --- Step 1: Handle Renaming (Unlink/Rename Parent) ---
    if "tank_number" in data:
        new_tank_number = data["tank_number"]
        if new_tank_number != tank.tank_number:
            # Check for duplicates before changing
            existing = db.query(Tank).filter(Tank.tank_number == new_tank_number).first()
            if existing:
                raise HTTPException(status_code=400, detail=f"Tank number '{new_tank_number}' already exists")
        
            # 1a. UNLINK: Temporarily set child FK to None to free up the parent
            tank_detail.tank_number = None
            
            # 1b. RENAME: Update the parent
            tank.tank_number = new_tank_number
            
            # --- Commit 1: Commit the parent update and the child unlink ---
            db.commit()

            # 1c. RELINK: Update the child to the new name (this is committed in the final step)
            tank_detail.tank_number = new_tank_number
    
    if "updated_by" in data:
        tank.updated_by = data["updated_by"]
    
    if "status" in data:
        status = data["status"]
        if status not in ["active", "inactive"]:
            raise HTTPException(status_code=400, detail="status must be 'active' or 'inactive'")
        tank.status = status # Sync status to parent

    # --- Step 3: Update Child Table (tank_details) ---
    detail_fields = [
        "status", "mfgr", "date_mfg", "pv_code", "un_iso_code",
        "capacity_l", "mawp", "design_temperature", "tare_weight_kg",
        "mgw_kg", "mpl_kg", "size", "pump_type",
        "vesmat", "gross_kg", "net_kg", "color_body_frame",
        "remark", "lease", "updated_by",
        # --- NEW FIELDS ---
        "working_pressure", "cabinet_type", "frame_type"
        # --------------------
    ]

    for field in detail_fields:
        if field in data:
            value = data[field]
            
            if field in ["capacity_l", "mawp", "tare_weight_kg", "mgw_kg", "mpl_kg", "gross_kg", "net_kg", "working_pressure"]:
                setattr(tank_detail, field, to_float_or_none(value))
            elif field == "lease":
                setattr(tank_detail, field, bool(value))
            elif field == "date_mfg":
                setattr(tank_detail, field, value or None) # Convert "" to None
            else:
                setattr(tank_detail, field, value)

    # --- Final Commit for both tables ---
    try:
        db.commit()
        db.refresh(tank)
        db.refresh(tank_detail)
    except Exception as e:
        # If the final commit fails, roll back everything
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update tank details: {str(e)}")
    
    return {"message": "Tank updated successfully"}


@router.delete("/{tank_id}")
def delete_tank(tank_id: int, db: Session = Depends(get_db)):
    tank_detail = db.query(TankDetails).filter(TankDetails.tank_id == tank_id).first()
    tank = db.query(Tank).filter(Tank.id == tank_id).first()

    if not tank:
        raise HTTPException(status_code=404, detail="Tank not found")

    if tank_detail:
        db.delete(tank_detail)
    db.delete(tank)
    db.commit()

    return {"message": "Tank deleted successfully"}

@router.get("/{tank_id}")
def get_tank_by_id(tank_id: int, db: Session = Depends(get_db)):
    tank = db.query(Tank).filter(Tank.id == tank_id).first()
    tank_detail = db.query(TankDetails).filter(TankDetails.tank_id == tank_id).first()

    if not tank or not tank_detail:
        raise HTTPException(status_code=404, detail="Tank not found")

    return {
        "id": tank.id,
        "tank_number": tank.tank_number,
        "status": tank_detail.status,
        "mfgr": tank_detail.mfgr,
        "date_mfg": tank_detail.date_mfg,
        "pv_code": tank_detail.pv_code,
        "un_iso_code": tank_detail.un_iso_code,
        "capacity_l": tank_detail.capacity_l,
        "mawp": tank_detail.mawp,
        "design_temperature": tank_detail.design_temperature,
        "tare_weight_kg": tank_detail.tare_weight_kg,
        "mgw_kg": tank_detail.mgw_kg,
        "mpl_kg": tank_detail.mpl_kg,
        "size": tank_detail.size,
        "pump_type": tank_detail.pump_type,
        "vesmat": tank_detail.vesmat,
        "gross_kg": tank_detail.gross_kg,
        "net_kg": tank_detail.net_kg,
        "color_body_frame": tank_detail.color_body_frame,
        "remark": tank_detail.remark,
        "lease": int(tank_detail.lease),
        "created_by": tank.created_by,
        "updated_by": tank.updated_by,
        # --- NEW FIELDS ---
        "working_pressure": tank_detail.working_pressure,
        "cabinet_type": tank_detail.cabinet_type,
        "frame_type": tank_detail.frame_type,
        # ------------------
    }

@router.get("/by-number/{tank_number}")
def get_tank_by_number(tank_number: str, db: Session = Depends(get_db)):
    tank = db.query(Tank).filter(Tank.tank_number == tank_number).first()
    if not tank:
        raise HTTPException(status_code=404, detail="Tank not found")

    tank_detail = db.query(TankDetails).filter(TankDetails.tank_id == tank.id).first()
    if not tank_detail:
        raise HTTPException(status_code=404, detail="Tank details not found")

    return {
        "tank_number": tank.tank_number,
        "date_mfg": tank_detail.date_mfg
    }